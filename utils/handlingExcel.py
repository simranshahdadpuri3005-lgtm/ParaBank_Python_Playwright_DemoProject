# pip install openpyxl
from openpyxl import load_workbook


def read_excel_by_rows(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)    
    return data
    # iter_rows is a method that returns the rows of the sheet as tuples
    # min_row=2 is used to skip the header row
    # values_only=True is used to get only the cell values without any formatting

    #iter_cols can be used to iterate through columns

    # firstRowFirstColumnValue = sheet.cell(row=2, column=1).value
    # firstRowSecondColumnValue = sheet.cell(row=2, column=2).value
    # return firstRowFirstColumnValue, firstRowSecondColumnValue  

def read_excel_by_cols(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_cols(min_col=1, values_only=True):
        data.append(row)    
    return data

def addData_excel(file_path, sheet_name, rowNo, columnNo, value):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=rowNo, column=columnNo).value = value
    workbook.save(file_path)

def append_to_excel(file_path, sheet_name, headers, rows):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(file_path)